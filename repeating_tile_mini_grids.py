from itertools import product
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class Minigrid:
    x_length: int
    y_length: int
    minigrid: list[str]

    def __init__(self, x_length, y_length, vals=[]):
        # initiates an EMPTY minigrid
        self.x_length = x_length
        self.y_length = y_length
        self.minigrid = vals

    def get_all_perms(self) -> list:
        combs = product('WB', repeat=(self.x_length * self.y_length))
        result = []
        for comb in combs:
            result.append([letter for letter in comb])
        print("Permutations complete")
        print(len(result))
        return result

    def to_big_grid(self) -> list[list[str]]:
        temp = []
        for i in range(self.y_length):
            row = []
            for j in range(self.x_length * 2):
                row.append(self.minigrid[(j % self.x_length) + (self.x_length * i)])
            temp.append(row)
        result = []
        for row in temp:
            result.append(row)
        for row in temp:
            result.append(row)
        return result

    def follows_rules(self) -> bool:
        """
        >>> grid1 = 'BWWWWWWBBWWWWWBWWWWBBBWWW'
        >>> mini1 = Minigrid(5,5, [char for char in grid1])
        >>> mini1.follows_rules()
        True
        >>> grid2 = 'WWWBBWWBWWWBWWWWBWWWBBWWW'
        >>> mini2 = Minigrid(5,5, [char for char in grid2])
        >>> mini2.follows_rules()
        False
        >>> grid3 = 'BWWBBWWBBBWBWWWWBWWWBBBBB'
        >>> mini3 = Minigrid(5,5, [char for char in grid3])
        >>> mini3.follows_rules()
        False
        >>> grid4 = 'BWWWWWBWWWWWBWWWWWBWWWWWB'
        >>> mini4 = Minigrid(5,5, [char for char in grid4])
        >>> mini4.follows_rules()
        True
        >>> grid5 = 'WWWWWWBWWWWWBWWWWWBWWWWWB'
        >>> mini5 = Minigrid(5,5, [char for char in grid5])
        >>> mini5.follows_rules()
        False
        """
        big_grid = self.to_big_grid()
        # across
        for i in range(len(big_grid)):
            last_black = -1
            for j in range(len(big_grid[0])):
                if big_grid[i][j] == 'B':
                    if last_black >= 0 and (j - last_black == 2 or j - last_black == 3):
                        return False
                    last_black = j
            if last_black == -1:
                return False
        # down
        for j in range(len(big_grid[0])):
            last_black = -1
            for i in range(len(big_grid)):
                if big_grid[i][j] == 'B':
                    if last_black >= 0 and (i - last_black == 2 or i - last_black == 3):
                        return False
                    last_black = i
            if last_black == -1:
                return False
        return True


if __name__ == "__main__":
    xlen = 5
    ylen = 5
    my_grid = Minigrid(xlen, ylen)
    perms = my_grid.get_all_perms()
    minigrids = []
    for perm in perms:
        temp_grid = Minigrid(xlen, ylen, perm)
        minigrids.append(temp_grid)
    print("List of minigrids complete")
    wb = Workbook()
    count = 1
    for grid in minigrids:
        big_grid = grid.to_big_grid()
        cols = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        if grid.follows_rules():
            wb.create_sheet(f'sheet_{count}')
            for i in range(len(big_grid)):
                for j in range(len(big_grid[0])):
                    cell = big_grid[i][j]
                    if cell == 'B':
                        wb[f'sheet_{count}'][f'{cols[j]}{i + 1}'].fill = PatternFill(
                            start_color="000000", end_color="000000", fill_type="solid")
            count += 1
    wb.save('grids.xlsx')
    print("Done")
